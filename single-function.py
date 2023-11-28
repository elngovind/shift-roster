from openpyxl import Workbook
from openpyxl.styles import Border, Side
import calendar
import pandas as pd
import numpy as np

def generate_shift_roster():
    # Collect user input
    days = int(input("Enter the number of days: "))
    resource_names = input("Enter resource names separated by space: ").split()

    # Define the shift types
    shift_types = ['M', 'A', 'N']

    # Create a list to store the shift data
    shifts_data = []

    # Generate shifts for each day and resource
    for day in range(1, days + 1):
        shifts = np.random.permutation(shift_types).tolist()
        for i, resource in enumerate(resource_names):
            shift = shifts[i % len(shifts)]
            shifts_data.append({'Date': day, 'Day_Name': calendar.day_name[(day - 1) % 7],
                                'Resource': resource, 'Shift_Type': shift})
            shifts.remove(shift)

    # Create a DataFrame from the collected data
    shifts_df = pd.DataFrame(shifts_data)
    shifts_pivot = shifts_df.pivot_table(index=['Date', 'Day_Name'], columns='Resource', values='Shift_Type', aggfunc='first')

    # Create an Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'Shift Roster'

    # Add headers for Dates and Days
    date_column = [f"Date_{day:02d}" for day in range(1, days + 1)]
    day_column = [calendar.day_name[(day - 1) % 7] for day in range(1, days + 1)]

    ws.append([''] + date_column)
    ws.append([''] + day_column)

    # Add the shift data to the worksheet
    for resource in resource_names:
        row = [resource]
        for day in range(1, days + 1):
            shift_code = shifts_pivot.loc[(day, calendar.day_name[(day - 1) % 7]), resource]
            row.append(shift_code if not pd.isnull(shift_code) else '')
        ws.append(row)

    # Add borders to the worksheet
    max_row = ws.max_row
    max_col = ws.max_column
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if cell.row == 1 or cell.column == 1:
                continue
            cell.border = border

    # Save the Excel file
    file_name = f'shift_roster_{days}_days.xlsx'
    wb.save(file_name)
    print(f"Shift roster generated and saved to '{file_name}'")

# Execute the function
generate_shift_roster()
