from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill
import calendar
import numpy as np
import pandas as pd

def generate_shift_roster(num_days, resource_names):
    # Define the resources
    resources = resource_names

    # Create a list to store the shift data
    shifts_data = []

    # Assign shifts to resources ensuring each day has all three shifts covered by different resources
    shift_types = ['M', 'A', 'N']
    for day in range(1, num_days + 1):
        np.random.shuffle(resources)
        shifts = shift_types.copy()
        for resource in resources:
            shift = shifts.pop(0)
            shifts_data.append({'Date': day, 'Day_Name': calendar.day_name[(day - 1) % 7],
                                'Resource': resource, 'Shift_Type': shift})
            if shift == 'M':
                shifts.append('O')  # Fill remaining shifts with 'O'

    # Create a DataFrame from the collected data
    shifts_df = pd.DataFrame(shifts_data)

    # Pivot the DataFrame to have dates as rows and resources in columns
    shifts_pivot = shifts_df.pivot_table(index=['Date', 'Day_Name'], columns='Resource', values='Shift_Type', aggfunc='first')

    # Assign 2 days off for each person per week
    for resource in resources:
        resource_schedule = shifts_pivot[resource]
        off_days = resource_schedule[resource_schedule == 'O'].groupby('Date').size()
        available_off_days = off_days[off_days < 2].index.tolist()
        for day in available_off_days:
            shifts_pivot.loc[day, resource] = 'O'

    # Create a new Workbook from openpyxl
    wb = Workbook()

    # Create a worksheet for the shift roster
    ws = wb.active
    ws.title = 'Shift Roster'

    # Add headers for Dates and Days
    date_column = [f"Date_{day:02d}" for day in range(1, num_days + 1)]
    day_column = [calendar.day_name[(day - 1) % 7] for day in range(1, num_days + 1)]

    ws.append([''] + date_column)
    ws.append([''] + day_column)

    # Add the shift abbreviations corresponding to resources and dates
    for resource in resources:
        row = [resource]
        for day in range(1, num_days + 1):
            shift_code = shifts_pivot.loc[(day, calendar.day_name[(day - 1) % 7]), resource]
            row.append(shift_code if not pd.isnull(shift_code) else '')
        ws.append(row)

    # Add border to the outside of the data
    max_row = ws.max_row
    max_col = ws.max_column
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if cell.row == 1 or cell.column == 1:
                continue
            cell.border = border

    # Save the Excel file
    file_name = f'shift_roster_{num_days}_days.xlsx'
    wb.save(file_name)
    print(f"Shift roster generated and saved to '{file_name}'")

# Input number of days and resource names during runtime
days = int(input("Enter the number of days: "))
resource_names = input("Enter resource names separated by space: ").split()
generate_shift_roster(days, resource_names)
